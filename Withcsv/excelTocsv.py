#! PYTHON 3
#! CONVERT all EXCEL File int he WORKING directory to csv file and sheet

import openpyxl, os,csv

for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):
        continue

    wb = openpyxl.load_workbook(excelFile)
    baseName, extention = excelFile.split('.')

    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheetName]
        newFileName = baseName + '_' + sheetName
        # Create the CSV filename from the Excel filename and sheet title.
        # Create the csv.writer object for this CSV file.
        csvFile = open(newFileName + '.csv', 'w', newline='')
        outputWriter = csv.writer(csvFile)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)
            outputWriter.writerow(rowData)
            # Write the rowData list to the CSV file.
