import openpyxl
wb = openpyxl.Workbook() # Create a blank workbook.

print(wb.sheetnames)

sheet = wb.active
print(sheet.title)

sheet.title = 'Spam Bacon Eggs Sheet' #Change title.
print(wb.sheetnames)

sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx') # Save the workbook

#sheets can be added to and removed from a Workbook

wb.create_sheet() #add a new sheet.
print(wb.sheetnames)
['Sheet','Sheet1']

#Create a new sheet at index 0.
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)
['First Sheet', 'Sheet', 'Sheet 1']

wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)
['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet 1']

#deleting sheets by title
del wb['Middle Sheet']
del wb['Sheet1']
print(wb.sheetnames)
['First Sheet', 'Sheet']

#Writing values to cells

sheet['A1'] = 'Hello, world' # Edit the cell's value.
print(sheet['A1']).values
