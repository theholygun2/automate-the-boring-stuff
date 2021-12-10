import openpyxl
import sys
from openpyxl.utils import get_column_letter


#switch values of row and column
wb = openpyxl.load_workbook('yogsz.xlsx')
sheet = wb.active
wbw = openpyxl.Workbook()
sheet2 = wbw.active

for y in range(1, sheet.max_row +1):
    for x in range(1, sheet.max_column +1):
        sheet2.cell(x, y).value = sheet.cell(y,x).value


wbw.save('takiz.xlsx')
