import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import sys

wb = openpyxl.Workbook()
sheet = wb.active

inputList = sys.argv
styleObj = Font(bold=True)
inputz = inputList[1]

# B1 = 1, C1 = 2, B2 = A2 * B1
# A2 = 1, A3 = 2, B3 = A3 * B1
[' ','B1', 'C1']
['A2','B2','C2']
['A3','B3','C3']

for x in range(1, int(inputz) + 1):
    sheet.cell(row=1, column=x+1).value = x
    sheet.cell(row=1, column=x+1).font = styleObj

for y in range(1, int(inputz) + 1):
    sheet.cell(row=y+1, column=1).value = y
    sheet.cell(row=y+1, column=1).font = styleObj

for rowx in range(2, int(inputz) + 2):
    for columny in range(2, int(inputz) + 2):
        var1 = 'A' + str(rowx)
        var2 = get_column_letter(columny) + str(1)
        sheet.cell(row=rowx, column=columny).value = f'={var1}*{var2}'
wb.save('multiplytable.xlsx')
