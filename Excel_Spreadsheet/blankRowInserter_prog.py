import openpyxl
import sys
from openpyxl.utils import get_column_letter

argsList = sys.argv
N = int(argsList[1])
M = int(argsList[2])
fileLoc = argsList[3]


wb = openpyxl.load_workbook(fileLoc)
sheet = wb.active
max_column = sheet.max_column

wb2 = openpyxl.Workbook('done.xlsx')
sheet2 = wb2.active


for i in range(N, N + M):
    for y in range(1, max_column + 1):
        sheet[f'{get_column_letter(y)}' + str(i)].value = None

#wb.save('updatedProduceSales.xlsx')
