import openpyxl

wb = openpyxl.Workbook()
sheet = wb['Sheet']

# FORMULAS
sheet['B9'] = 'SUM(B1:B8)'
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = 'SUM(A1:A2)' #Set the Formula

#Adjusting Rows and Columns

#Setting Row Height and Column Width

sheet['A1'] = 'Tall row'
sheet['B1'] = 'Wide column'

#set the height and Width height = 0-409, width = 0-255
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].Width = 20

#MERGING and Unmerging cells

sheet.merge_cells('A1:D3') #Merge all these cells.
sheet['A1'] = "Twelve cells merged together"
sheet.merge_cells('C5:D5')#merge two cells
sheet['C5'] = 'Two merged cells'

sheet.unmerge_cells('A1:D3') # Split these cells up.
sheet.unmerge_cells('C5:D5')
