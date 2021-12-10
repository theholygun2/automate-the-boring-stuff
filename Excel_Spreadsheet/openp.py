import openpyxl as xl
from openpyxl.utils import get_column_letter, column_index_from_string

#opening Excel Documents using openpyxl
wb = xl.load_workbook('example.xlsx')
print(wb)
print(wb.sheetnames)

'''getting Sheets from the Workbook'''
sheet = wb['Sheet1']
print(sheet)

print(sheet.title)#get the sheet title as a string

anotherSheet = wb.active#get the active sheet output: <worksheet "Sheet1"
print(anotherSheet)

'''Getting cells from the Sheets'''

print(sheet['A1'])#get a cell from the sheet 3
print(sheet['A1'].value)#get the value from the cell
c = sheet['B1']#get another cell fromo the sheet
print(c.value)

'''Get the row, column, and value from the cell'''

print('Row %s, Column %s is %s' %(c.row, c.column, c.value))
print('Cell %s is %s' %(c.coordinate, c.value))
print(sheet['C1'].value)

print('sheet cell[row bla blabla]')
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1,8,2): #Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_row) #Get the
print(sheet.max_column) #Get the highest column number output integer

'''Converting Between Column Lettters and Numbers'''

print(get_column_letter(1))
print(get_column_letter(27))
print(get_column_letter(900))

print(get_column_letter(sheet.max_column))
print(column_index_from_string('A')) #Get A's Numbers
print(column_index_from_string('AA'))

print(tuple(sheet['A1':'C3'])) #Get all cells from A1 to C3
#((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>,<Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--end of row--')

print(list(sheet.columns)[1])#Get second columns's rowOfCellObjects
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
